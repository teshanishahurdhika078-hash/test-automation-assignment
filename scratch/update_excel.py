import openpyxl
import re
import random

wb = openpyxl.load_workbook('Assignment 1 - Test cases.xlsx')
ws = wb.active

for row in range(2, ws.max_row + 1):
    input_val = ws.cell(row=row, column=3).value
    expected_val = ws.cell(row=row, column=4).value
    rationale = ws.cell(row=row, column=9).value
    accuracy = ws.cell(row=row, column=7).value

    # Update Input
    if input_val and isinstance(input_val, str):
        # Remove stray ' n '
        input_val = re.sub(r'\b n\b\s*', '', input_val)
        
        # Replace 'bn' or 'Bn'
        if 'bn' in input_val.lower():
            if random.choice([True, False]):
                input_val = re.sub(r'\bbn\b', 'mcn', input_val)
                input_val = re.sub(r'\bBn\b', 'Mcn', input_val)
            else:
                # Just remove it
                input_val = re.sub(r'\bbn\b\s*', '', input_val, flags=re.IGNORECASE)
                
        # Randomize punctuation slightly
        if '???' in input_val:
            input_val = input_val.replace('???', '?!')
        elif '??' in input_val:
            input_val = input_val.replace('??', '?')
            
        ws.cell(row=row, column=3).value = input_val.strip()

    # Update Expected Output
    if expected_val and isinstance(expected_val, str):
        expected_val = expected_val.replace(' න්', '')
        if 'mcn' in (input_val or '').lower() or 'Mcn' in (input_val or '').lower():
            expected_val = expected_val.replace('බන්', 'මචන්')
        else:
            expected_val = expected_val.replace(' බන්', '')
        
        # Randomize punctuation slightly to match input
        if '?!' in (input_val or ''):
            expected_val = expected_val.replace('???', '?!').replace('??', '?!').replace('?', '?!')
        elif '??' not in (input_val or '') and '?' in (input_val or ''):
            expected_val = expected_val.replace('???', '?').replace('??', '?')
            
        ws.cell(row=row, column=4).value = expected_val.strip()

    # Update Rationale
    if rationale and isinstance(rationale, str):
        r = rationale.replace('Rationale: ', '')
        r = r.replace('The input is phrased as a question', 'This sentence acts as an inquiry')
        r = r.replace('Uses', 'Contains')
        r = r.replace('Phrases like', 'Sentences such as')
        r = r.replace('A direct instruction', 'Provides a command')
        r = r.replace('A common greeting', 'This is a normal greeting')
        ws.cell(row=row, column=9).value = f"Justification: {r.strip()}"

    # Update Accuracy / Issue Type
    if accuracy and isinstance(accuracy, str):
        a = accuracy.replace('Spelling Error', 'Typo')
        a = a.replace('Extra Words', 'Unnecessary Slang')
        a = a.replace('Word Order Error', 'Grammar/Order Issue')
        ws.cell(row=row, column=7).value = a

wb.save('Assignment 1 - Test cases.xlsx')
print("Excel updated successfully!")
